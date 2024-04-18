from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import Workbook

# Setup Chrome options
chrome_options = Options()
chrome_options.add_argument("--headless")  # Run in headless mode

# Path to your ChromeDriver
# Specify the correct path to the Chrome WebDriver executable
service = Service(executable_path=r'C:\Users\pc\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe')

# Initialize the driver
driver = webdriver.Chrome(service=service, options=chrome_options)

# URL to scrape
url = "https://www.semrush.com/trending-websites/in/market-research"
driver.get(url)

# Function to scrape websites
def scrape_websites(driver):
    website_data = set()  # Use set to store unique rows
    websites = driver.find_elements(By.CSS_SELECTOR, "div.LRHbiboxtc64CSvZQUbp")
    for website in websites:
        name = website.find_element(By.CSS_SELECTOR, "a.___SBoxInline_1tphm_gg_").text.strip()
        full_url = website.find_element(By.CSS_SELECTOR, "a.___SBoxInline_1tphm_gg_").get_attribute('href')
        trimmed_name = name.split('.')[0]  # Remove domain
        trimmed_url = full_url.replace("https://www.semrush.com/website/", "").replace("/overview?source=trending-websites", "")
        website_data.add((trimmed_name, trimmed_url))  # Add the row as a tuple to the set
    return website_data

all_website_data = scrape_websites(driver)

# Find and click "Show more" button until it's no longer available
while True:
    try:
        show_more_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "div.dWpoM7T_pxot7qnmadYT button.o4Xw9479CpShI3uKU2Sd"))
        )
        driver.execute_script("arguments[0].click();", show_more_button)
        sleep(1)  # Wait for the page to load more items
        all_website_data.update(scrape_websites(driver))  # Update the set with new unique rows
    except:
        break  # If no more buttons are found or clickable, exit the loop

# Close the driver
driver.quit()

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.append(["Name", "URL"])

# Add data to the worksheet
for name, url in all_website_data:
    ws.append([name, url])

# Add hyperlinks to the URLs
for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
    if len(row) == 2:  # Check if the row has expected number of elements
        ws.cell(row=row[0], column=2).hyperlink = row[1]

# Save the workbook
wb.save("market-research.xlsx")

print("Excel file has been created successfully!")
